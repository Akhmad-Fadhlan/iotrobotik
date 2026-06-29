import os
import json

def extract_excel_data():
    file_path = "Kumpulan Link Modul robotic iot IDN (1).xlsx"
    if not os.path.exists(file_path):
        print(f"Error: {file_path} not found.")
        return

    try:
        import openpyxl
        print("Using openpyxl...")
        wb = openpyxl.load_workbook(file_path)
        data = {}
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            rows = []
            for row in sheet.iter_rows(values_only=True):
                # Only keep rows that have some content
                if any(row):
                    rows.append([str(cell) if cell is not None else "" for cell in row])
            data[sheet_name] = rows
        
        output_file = "scripts/excel_data.json"
        with open(output_file, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2)
        print(f"Excel data successfully extracted to {output_file} using openpyxl!")
        return
    except ImportError:
        print("openpyxl is not installed. Trying zipfile/xml fallback...")

    # Fallback to pure python zipfile XML reading
    try:
        import zipfile
        import xml.etree.ElementTree as ET
        
        data = {}
        with zipfile.ZipFile(file_path) as z:
            # 1. Read workbook.xml to get sheet names
            wb_xml = z.read("xl/workbook.xml")
            root_wb = ET.fromstring(wb_xml)
            # Find namespaces
            ns = {'ns': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main',
                  'r': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'}
            
            sheets_info = []
            for sheet_el in root_wb.findall('.//ns:sheet', ns):
                name = sheet_el.attrib.get('name')
                sheet_id = sheet_el.attrib.get('sheetId')
                r_id = sheet_el.attrib.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id')
                sheets_info.append((name, r_id))
            
            # Read sharedStrings.xml
            shared_strings = []
            if "xl/sharedStrings.xml" in z.namelist():
                ss_xml = z.read("xl/sharedStrings.xml")
                root_ss = ET.fromstring(ss_xml)
                for t in root_ss.findall('.//ns:t', ns):
                    shared_strings.append(t.text or "")
            
            # 2. Read sheet XMLs
            # Map relationship ID to sheet filename
            # Read xl/_rels/workbook.xml.rels
            rels_xml = z.read("xl/_rels/workbook.xml.rels")
            root_rels = ET.fromstring(rels_xml)
            rel_ns = {'rels': 'http://schemas.openxmlformats.org/package/2006/relationships'}
            rel_map = {}
            for rel in root_rels.findall('.//rels:Relationship', rel_ns):
                r_id = rel.attrib.get('Id')
                target = rel.attrib.get('Target')
                rel_map[r_id] = target
            
            for name, r_id in sheets_info:
                target_path = rel_map.get(r_id)
                if target_path:
                    # Clean target path (sometimes starts with / or workbook/)
                    if target_path.startswith('/'):
                        target_path = target_path[1:]
                    if not target_path.startswith('xl/'):
                        target_path = 'xl/' + target_path
                    
                    if target_path in z.namelist():
                        sheet_xml = z.read(target_path)
                        root_sheet = ET.fromstring(sheet_xml)
                        
                        # Parse rows
                        rows_dict = {}
                        for r_el in root_sheet.findall('.//ns:row', ns):
                            r_num = int(r_el.attrib.get('r', 1)) - 1
                            row_cells = []
                            for c_el in r_el.findall('./ns:c', ns):
                                val_el = c_el.find('./ns:v', ns)
                                cell_val = ""
                                if val_el is not None:
                                    t_attrib = c_el.attrib.get('t')
                                    if t_attrib == 's':  # shared string
                                        idx = int(val_el.text)
                                        cell_val = shared_strings[idx] if idx < len(shared_strings) else ""
                                    else:
                                        cell_val = val_el.text or ""
                                row_cells.append(cell_val)
                            if any(row_cells):
                                rows_dict[r_num] = row_cells
                        
                        # Convert dict to sorted rows
                        max_row = max(rows_dict.keys()) if rows_dict else 0
                        sheet_data = []
                        for i in range(max_row + 1):
                            if i in rows_dict:
                                sheet_data.append(rows_dict[i])
                        data[name] = sheet_data

        output_file = "scripts/excel_data.json"
        with open(output_file, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2)
        print(f"Excel data successfully extracted to {output_file} using XML parsing!")
    except Exception as e:
        print(f"XML Parsing failed: {e}")

if __name__ == '__main__':
    extract_excel_data()
